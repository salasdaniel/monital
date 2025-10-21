from django.http import JsonResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views import View
from ..models import User, Matricula, Empresa
import jwt, json
from decouple import config
from .views import validate_app_key, validate_jwt
from django.shortcuts import get_object_or_404
from openpyxl import load_workbook
from io import BytesIO


@method_decorator(csrf_exempt, name='dispatch')
class MatriculaAddView(View):
    def post(self, request):
        # Validar app_key por seguridad
        invalid = validate_app_key(request)
        if invalid:
            return invalid

        # Validar JWT del usuario autenticado
        user, error_response = validate_jwt(request)
        if error_response:
            return error_response

        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        # Validar campos requeridos
        required_fields = ['nro_matricula']
        missing_fields = [field for field in required_fields if not data.get(field)]
        
        if missing_fields:
            return JsonResponse({
                'error': f'Campos requeridos faltantes: {", ".join(missing_fields)}'
            }, status=400)

        # Validar que el número de matrícula no exista
        if Matricula.objects.filter(nro_matricula=data.get('nro_matricula')).exists():
            return JsonResponse({'error': 'Ya existe una matrícula con este número'}, status=400)

        try:
            # Crear la nueva matrícula
            matricula = Matricula.objects.create(
                nro_matricula=data.get('nro_matricula'),
                tracker_id=data.get('tracker_id'),
                empresa_id=data.get('empresa_id'),
                usuario_creacion=user
            )

            return JsonResponse({
                'message': 'Matrícula creada exitosamente',
                'matricula': {
                    'id': matricula.id,
                    'nro_matricula': matricula.nro_matricula,
                    'tracker_id': matricula.tracker_id,
                    'empresa_id': matricula.empresa_id,
                    'empresa_nombre': matricula.empresa.nombre_comercial if matricula.empresa else None,
                    'created_at': matricula.created_at.isoformat(),
                    'usuario_creacion': matricula.usuario_creacion.username if matricula.usuario_creacion else None
                }
            }, status=201)

        except Exception as e:
            return JsonResponse({'error': f'Error al crear la matrícula: {str(e)}'}, status=500)

        
@method_decorator(csrf_exempt, name='dispatch')
class MatriculaUpdateView(View):
    def put(self, request, pk):
        # Validar app_key por seguridad
        invalid = validate_app_key(request)
        if invalid:
            return invalid

        # Validar JWT del usuario autenticado
        user, error_response = validate_jwt(request)
        if error_response:
            return error_response

        try:
            data = json.loads(request.body.decode('utf-8'))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'JSON inválido'}, status=400)

        try:
            # Obtener la matrícula
            matricula = get_object_or_404(Matricula, id=pk)
            
            # Solo se puede actualizar el tracker_id
            if 'tracker_id' in data:
                matricula.tracker_id = data.get('tracker_id')
                matricula.save()
            else:
                return JsonResponse({
                    'error': 'No se proporcionó el campo tracker_id para actualizar'
                }, status=400)

            return JsonResponse({
                'message': 'Matrícula actualizada exitosamente',
                'matricula': {
                    'id': matricula.id,
                    'nro_matricula': matricula.nro_matricula,
                    'tracker_id': matricula.tracker_id,
                    'empresa_id': matricula.empresa_id,
                    'empresa_nombre': matricula.empresa.nombre_comercial if matricula.empresa else None,
                    'created_at': matricula.created_at.isoformat(),
                    'updated_at': matricula.updated_at.isoformat(),
                    'usuario_creacion': matricula.usuario_creacion.username if matricula.usuario_creacion else None
                }
            }, status=200)
        
        except Http404:
            return JsonResponse({'error': 'La matrícula no existe'}, status=404)

        except Exception as e:
            return JsonResponse({'error': f'Error al actualizar la matrícula: {str(e)}'}, status=500)


class MatriculaListView(View):
    def get(self, request):
        # Validar app_key por seguridad
        invalid = validate_app_key(request)
        if invalid:
            return invalid

        # Validar JWT del usuario autenticado
        user, error_response = validate_jwt(request)
        if error_response:
            return error_response

        try:
            # Obtener el parámetro opcional empresa_id desde query string
            empresa_id = request.GET.get('empresa_id')
            
            # Filtrar matrículas
            if empresa_id:
                # Validar que la empresa exista
                try:
                    empresa = Empresa.objects.get(id=empresa_id)
                    matriculas = Matricula.objects.filter(empresa=empresa).order_by('-created_at')
                except Empresa.DoesNotExist:
                    return JsonResponse({
                        'error': 'La empresa especificada no existe'
                    }, status=400)
            else:
                # Obtener todas las matrículas
                matriculas = Matricula.objects.all().order_by('-created_at')
            
            # Serializar los datos
            matriculas_data = []
            for matricula in matriculas:
                matriculas_data.append({
                    'id': matricula.id,
                    'nro_matricula': matricula.nro_matricula,
                    'tracker_id': matricula.tracker_id,
                    'empresa_id': matricula.empresa_id,
                    'empresa_nombre': matricula.empresa.nombre_comercial if matricula.empresa else None,
                    'created_at': matricula.created_at.isoformat(),
                    'updated_at': matricula.updated_at.isoformat(),
                    'usuario_creacion': matricula.usuario_creacion.username if matricula.usuario_creacion else None
                })

            return JsonResponse({
                'message': 'Matrículas obtenidas exitosamente',
                'count': len(matriculas_data),
                'matriculas': matriculas_data
            }, status=200)

        except Exception as e:
            return JsonResponse({'error': f'Error al obtener las matrículas: {str(e)}'}, status=500)


@method_decorator(csrf_exempt, name='dispatch')
class MatriculaImportView(View):
    def post(self, request):
        # Validar app_key por seguridad
        invalid = validate_app_key(request)
        if invalid:
            return invalid

        # Validar JWT del usuario autenticado
        user, error_response = validate_jwt(request)
        if error_response:
            return error_response

        # Validar que se haya enviado un archivo
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No se proporcionó ningún archivo'}, status=400)

        file = request.FILES['file']

        # Validar la extensión del archivo
        if not file.name.endswith('.xlsx'):
            return JsonResponse({'error': 'El archivo debe ser un archivo XLSX'}, status=400)

        try:
            # Leer el archivo Excel
            workbook = load_workbook(filename=BytesIO(file.read()))
            sheet = workbook.active

            # Validar que el archivo tenga al menos 2 columnas
            if sheet.max_column < 2:
                return JsonResponse({
                    'error': 'El archivo debe tener al menos 2 columnas (matricula, tracker_id)'
                }, status=400)

            # Contadores para el resultado
            created_count = 0
            updated_count = 0
            errors = []

            # Procesar cada fila (saltando la primera fila de encabezados)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Obtener valores de las columnas
                    nro_matricula = str(row[0]).strip() if row[0] else None
                    tracker_id = str(row[1]).strip() if row[1] else None

                    # Validar que la matrícula no esté vacía
                    if not nro_matricula:
                        errors.append({
                            'fila': row_num,
                            'error': 'Número de matrícula vacío'
                        })
                        continue

                    # Buscar si la matrícula existe
                    matricula = Matricula.objects.filter(nro_matricula=nro_matricula).first()

                    if matricula:
                        # Si existe, actualizar el tracker_id
                        matricula.tracker_id = tracker_id
                        matricula.save()
                        updated_count += 1
                    else:
                        # Si no existe, crear nueva matrícula
                        Matricula.objects.create(
                            nro_matricula=nro_matricula,
                            tracker_id=tracker_id,
                            usuario_creacion=user
                        )
                        created_count += 1

                except Exception as e:
                    errors.append({
                        'fila': row_num,
                        'error': str(e)
                    })

            return JsonResponse({
                'message': 'Importación completada',
                'created': created_count,
                'updated': updated_count,
                'errors': errors,
                'total_processed': created_count + updated_count,
                'total_errors': len(errors)
            }, status=200)

        except Exception as e:
            return JsonResponse({'error': f'Error al procesar el archivo: {str(e)}'}, status=500)
